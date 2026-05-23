# Lifted from project-mcp-server/parsers/excel_parser.py per M0-06 audit (REUSE bucket).
# Adapted: project_id -> tenant_id + source_id; removed direct DB writes
# (results return as Pydantic models, persistence is the ingestion service's job).
# Also dropped the `ScheduleExcelParser` AEC-specific subclass per the audit's
# "lift wholesale minus the subclass" directive (M1-ING-03 LLM classifier
# replaces those subclasses).
"""Excel Parser - Extracts data from spreadsheets.

Handles:
  - .xlsx and .xls files
  - .csv files
  - Multiple sheets
  - Tabular data extraction
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Optional

from .base import BaseParser


class ExcelParser(BaseParser):
    document_type = "general_document"
    supported_extensions = [".xlsx", ".xls", ".csv"]
    supported_mime_types = [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
        "text/csv",
    ]

    def extract_text(self, file_path: Path) -> str:
        """Convert spreadsheet to searchable text."""
        if file_path.suffix.lower() == ".csv":
            return self._extract_csv(file_path)

        import openpyxl

        wb = openpyxl.load_workbook(str(file_path), data_only=True)
        parts = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"\n=== Sheet: {sheet_name} ===\n")

            for row in ws.iter_rows(values_only=True):
                row_text = " | ".join(
                    str(cell) if cell is not None else "" for cell in row
                )
                if row_text.replace("|", "").strip():
                    parts.append(row_text)

        return "\n".join(parts)

    def _extract_csv(self, file_path: Path) -> str:
        """Extract text from CSV file."""
        import csv

        parts: list[str] = []
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            for row in reader:
                parts.append(" | ".join(row))
        return "\n".join(parts)

    def extract_fields(self, file_path: Path, full_text: str) -> dict[str, Any]:
        fields = {
            "title": file_path.stem,
        }
        return fields

    def get_sheet_names(self, file_path: Path) -> list[str]:
        """Get list of sheet names."""
        import openpyxl

        wb = openpyxl.load_workbook(str(file_path), data_only=True)
        return wb.sheetnames

    def get_sheet_as_dicts(
        self, file_path: Path, sheet_name: Optional[str] = None
    ) -> list[dict[str, Any]]:
        """Get a sheet as a list of dicts (first row = headers)."""
        import openpyxl

        wb = openpyxl.load_workbook(str(file_path), data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return []

        headers = [
            str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])
        ]
        return [
            {headers[i]: cell for i, cell in enumerate(row)}
            for row in rows[1:]
            if any(cell is not None for cell in row)
        ]
